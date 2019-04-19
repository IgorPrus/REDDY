# -*- coding: utf-8 -*-
'''
Created on 18.04.2019 by Dr.REDDY'S LABORATORIES LTD
@author: Igor Prus <igorbprus@gmail.com>
'''


from tkinter import *

from tkinter import filedialog as fd
from tkinter import messagebox as mb
import unicodedata
from openpyxl import load_workbook

# Модуль для работы с базой данных
import sqlite3 as sql
# Модуль для работы с COM объектами
import win32com.client
# Модуль создания Excel файлов
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# Модуль стилей создания стилей Excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# Модуль времени
import datetime
# Модуль создания папок 
import os
# Модуль для работы с файлами
import shutil
# Модуль чтение Excel файлов
import xlrd, xlwt
# Модуль преобразования чисел в слова
from num2words import num2words


	
def insertText():
	# определяем текущий каталог
	path_tmp = os.getcwd()
	path = path_tmp + "/tmp/"

	# Создаем католог по дате
	try:
		os.mkdir(path)
	except OSError:
		s = ""
		text.insert(END, s)
	else:
		s = "Успешно создана директория %s " % path + "\n\n"
		text.insert(END, s)
	try:
		file_name = fd.askopenfilename(filetypes = (("Excel", "*.xls"),
												("All files", "*.*") ))
		f = open(file_name)
		f.close()
		s = str(f)
		s = s + "\n\n"
		text.insert(1.0, s)
		f=str(f)
		f = f.replace("<_io.TextIOWrapper name=\'","")
		f = f.replace("' mode='r' encoding='cp1251'>","")
		s = str(f)
		s = s + "\n\n"
		text.insert(END, s)
		shutil.copyfile(f, 'tmp/tmp.xls')
		# s = f.read()
		# text.insert(1.0, s)
		#shutil.copyfile(r'tmp/1.xlsx', r'tmp/tmp/tmp.xlsx')
		#f.close()
	except FileNotFoundError:
		mb.showinfo("Внимание", "Файл не загружен")
	s = "Файл успешно загружен!!!\n\n"
	text.insert(END, s)

	
def preobrazovat():
	# Создание и подключение к базе данных
	connection = sql.connect('tmp/tmp.sqlite')
	# Создание курсора
	q = connection.cursor()
	q.execute('''DROP TABLE tmp ''')
	connection.commit()
	text.delete(1.0, END)
	s = "Очистка базы данных" + "\n\n"
	text.insert(END, s)
	
def virvat():
	try:
		file_name_xls = datetime.datetime.now()

		file_name_xls = str(file_name_xls)
		file_name_xls = file_name_xls.replace(':','_')
		file_name_xls = file_name_xls.replace('.','_')
		file_name_xls = file_name_xls.replace('-','_')
		print(file_name_xls)


		# определяем текущий каталог
		path_tmp = os.getcwd()
		path = path_tmp + "/tmp/" + file_name_xls

		# Создаем католог по дате
		try:
			os.mkdir(path)
		except OSError:
			s = "Создать директорию %s не удалось" % path + "\n\n"
			text.insert(END, s)
		else:
			s = "Успешно создана директория %s " % path + "\n\n"
			text.insert(END, s)


		#открываем файл
		rb = xlrd.open_workbook('tmp/tmp.xls',formatting_info=True)

		#выбираем активный лист
		sheet = rb.sheet_by_index(0)

		#получаем значение первой ячейки A1
		val = sheet.row_values(0)[0]

		#получаем список значений из всех записей

		vals = [sheet.row_values(rownum) for rownum in range(sheet.nrows)]


		i = 0
		while i <=100:
			try:
				vals.remove(['', '', '', '', '', '', '', '', ''])
				i+=1
			except:
				break


		vals.reverse()

		i = 0
		while i <=100:
			try:
				vals.remove(['', '', '', '', '', '', '', '', ''])
				i+=1
			except:
				break

		vals.reverse()

		vals.pop(0)

		i = 0
		while i <=100:
			try:
				vals.remove(['', '', '', '', '', '', '', '', ''])
				i+=1
			except:
				break

		vals.pop(1)
		vals.pop(1)
		vals.pop()
		vals.pop()
		#print(vals)
	

		#
		# Работа с базой данных
		#

		# Создание и подключение к базе данных
		connection = sql.connect('tmp/tmp.sqlite')
		# Создание курсора
		q = connection.cursor()

		q.execute('''CREATE TABLE tmp (id int auto_increment primary key, 
										GL_CODE varchar,
										description varchar,								
										date varchar,
										Invoice_No varchar,
										Value_in_Local_currency varchar,
										Rate varchar,
										Total_in_USD real,
										Total_in_BYR real)''')
		connection.commit()


		# Запись в базу данных
		i = 0
		while i<=1000:
			try:
				x = vals[i]
				code = x[0]
				description = x[1]
				date = x[2]
				invoice = x[3]
				value_in_local = x[4]
				rate = x[5]
				total_in_usd = x[6]
				total_in_byr = x[7]
				q.execute("INSERT INTO tmp (GL_CODE, description, date, Invoice_No, Value_in_Local_currency, Rate, Total_in_USD, Total_in_BYR) VALUES ('%s' , '%s', '%s', '%s', '%s', '%s', '%s', '%s')"%(code, description, date, invoice, value_in_local, rate, total_in_usd, total_in_byr))
				connection.commit()
				i+=1
			except:
				break
		




		q.execute("SELECT * FROM tmp")
		row = q.fetchone()



		while row is not None:

			wb = Workbook()

			# Создание ваучера
			dest_filename = 'tmp/'+file_name_xls+'/vaycher_#'+row[4]+'.xlsx'

			ws1 = wb.active

			ws1.title = "Лист 1"

			# Стиль ячеек D4
			ft = Font(name = 'Times New Roman', size = 12, bold = True, italic = True)
			ft1 = Font(name = 'Arial', size = 12)
			ft2 = Font(name = 'Arial', size = 10)

			ws1['D4'] =  'Dr.REDDY\'S LABORATORIES LTD'
			b4 = ws1['D4']
			b4.font = ft
			
			ws1['G7'] =  'Date/Дата __' +row[3] + '__'
			g7 = ws1['G7']
			g7.font = ft2

			ws1['D5'] =  'PAYMENT VOUCHER'
			b5 = ws1['D5']
			b5.font = ft
			
			a9 = 'Payee/Получатель ___Dhaipulle Sreenivasa Rao____________________________________________________________________________________________'
			a9 = a9[0:91]
			ws1['A9'] =  a9
			a9 = ws1['A9']
			a9.font = ft1

			
			a11 =  'Towards/ Цель_______' +row[2] + '_____________________________________________________________________________________________________________'
			a11 = a11[0:91]
			ws1['A11'] = a11
			a11 = ws1['A11']
			a11.font = ft1

			a13 =  '_________________________________________________________________________________________________________________________________________________________________________'
			a13 = a13[0:91]
			ws1['A13'] = a13
			a13 = ws1['A13']
			a13.font = ft1
	
	
			usd = row[7]
			try:
				usd = round(usd, 2)
			except TypeError:
				usd = usd
			byr = row[8]
			try:
				byr = round(byr, 2)
			except TypeError:
				byr = byr
			usd = str(usd)
			byr = str(byr)
	
			ws1['A15'] =  'Amount/сумма( в цифрах) (Rubles/Dollars) :________'+byr+' руб.коп. /'+usd+' USD'+'_________________________'
			a15 = ws1['A15']
			a15.font = ft1
	
	

			byr = row[8]
			if type(byr) == str:
				byr_str = '0'
			else:
				byr_str = num2words(byr, lang='ru', to = "currency")
				ws1['A19'] = 'Amount in words ( Сумма прописью ) _____'+ byr_str
				a19 = ws1['A19']
				a19.font = ft1
				
			usd = row[7]
			if type(usd) == str:
				usd_str = '0'
			else:
				usd_str = num2words(usd, lang='en', to = "currency")
				ws1['A21'] = '__________________________________'+ usd_str
				a21 = ws1['A21']
				a21.font = ft1

			ws1['A24'] =  'Issued by              Approved by             Executive Director               Received by/Получатель'
			a24 = ws1['A24']
			a24.font = ft1
			
			ws1['D27'] =  'FOR OFFICE USE ONLY'
			d27 = ws1['D27']
			d27.font = ft1
			
			ws1['A29'] =  'Voucher № _____'+row[4] +'__________' 
			a29 = ws1['A29']
			a29.font = ft1
			
			ws1['F29'] =  'Debit Account __________________'
			f29 = ws1['F29']
			f29.font = ft1
			
			ws1['A31'] =  'Conversion rate ____'+row[6]+'_____________' 
			a31 = ws1['A31']
			a31.font = ft1
			
			ws1['F31'] =  'Amount (Dollars) ____________________'
			f31 = ws1['F31']
			f31.font = ft1


			row = q.fetchone()
			wb.save(filename = dest_filename)
		

		# Удаление таблицы базы данных

		q.execute('''DROP TABLE tmp ''')
		connection.commit()
		s = "Очистка базы данных" + "\n\n"
		text.insert(END, s)
		# Отключение от базы данных
		q.close()
		connection.close()
		# Удаление файла
		os.remove("tmp/tmp.xls")
		s = "Удаление временных файлов" + "\n\n"
		text.insert(END, s)
		s = "Файлы доступны по адресу: %s " % path + "\n\n"
		text.insert(END, s)
		s = "Завершено !!!\n\n"
		text.insert(END, s)
		mb.showinfo("Ура", "Завершено!!!\n Файлы доступны по адресу: %s " % path )
	except FileNotFoundError:
		mb.showinfo("Внимание", "Файл не загружен")
		s = "Файл не загружен, пожулуйста загрузите файл и повторите!!!\n\n"
		text.insert(END, s)
	except:
		s = "База данных не пуста, пожалуйста очистите базу данный!!!\n\n"
		text.insert(END, s)
	
app = Tk()
app.title("Dr.REDDY'S LABORATORIES LTD")
#Название программы

mainmenu = Menu(app) 
app.config(menu=mainmenu) 
 
filemenu = Menu(mainmenu, tearoff=0)
filemenu.add_command(label="Открыть...")
filemenu.add_command(label="Новый")
filemenu.add_command(label="Сохранить...")
filemenu.add_command(label="Выход")
 
helpmenu = Menu(mainmenu, tearoff=0)
helpmenu.add_command(label="Помощь")
helpmenu.add_command(label="О программе")
 
mainmenu.add_cascade(label="Файл", menu=filemenu)
mainmenu.add_cascade(label="Справка", menu=helpmenu)

name = Label(app, text = "Программа для Dr.REDDY'S LABORATORIES LTD\n для разбития Excel файла")
name.config(font = ("Comic Sans MS", 16))
name.grid(row = 0, column = 0, columnspan = 3)
#имя программы

i_text = Label(app, text = "Введите первую строку: ")
i_text.config(font = (18))
i_text.grid(row = 1, column = 0, sticky = E, pady=10, padx=10)

i_entry = Entry(app, width = 25)
i_entry.grid(row = 1, column = 1, sticky  = W, pady=10, padx=10)

n_text = Label(app, text = "Введите посдеднюю строку: ")
n_text.config(font = (18))
n_text.grid(row = 2, column = 0, sticky = E, pady=10, padx=10)

n_entry = Entry(app, width = 25)
n_entry.grid(row = 2, column = 1, sticky  = W, pady=10, padx=10)


b1 = Button(app, text="Загрузить файл", command=insertText)
b1.config(font = (16))
b1.grid(row=3, column=0)


b4 = Button(app, text="Очистка данных", command=preobrazovat)
b4.config(font = (16))
b4.grid(row=4, column=2)

b5 = Button(app, text="Разбить файл", command=virvat)
b5.config(font = (16))
b5.grid(row=3, column=1)

text = Text(width=50, height=25)
text.grid(columnspan=3)
scroll = Scrollbar(command=text.yview)
# scroll.pack(side=LEFT, fill=Y)
 
text.config(yscrollcommand=scroll.set)
 
app.mainloop() 